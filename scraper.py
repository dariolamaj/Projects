import os
import requests
from datetime import date, timedelta
from openpyxl import load_workbook

yesterday = date.today() - timedelta(days=1)
today = date.today()

url = "https://transparency.entsog.eu/api/v1/operationalData.xlsx?from=" + str(yesterday) + "&indicator=Physical%20Flow&limit=-1&periodType=hour&periodize=0&pointDirection=sk-tso-0001itp-00117entry&to="+str(today)+ "&timezone=CET"
response = requests.get(url=url)
with open(str(today)+ ".xlsx", "wb") as f:
    
    f.write(response.content)
yesterday2 = date.today() - timedelta(days=2)
url = "https://transparency.entsog.eu/api/v1/operationalData.xlsx?from=" + str(yesterday2) + "&indicator=Physical%20Flow&limit=-1&periodType=hour&periodize=0&pointDirection=sk-tso-0001itp-00117entry&to="+str(yesterday)+ "&timezone=CET"
response2 = requests.get(url=url)
with open(str(yesterday)+ ".xlsx", "wb") as g:
    
    g.write(response2.content)
wb = load_workbook(str(today)+ ".xlsx")
wb.create_sheet(str(yesterday))
wb.save(str(today)+ ".xlsx")
wb2 = load_workbook(str(yesterday)+ ".xlsx")
wb2_sheet = wb2["page"]
wb_sheet = wb[str(yesterday)]
wb.save(str(today)+ ".xlsx")
for i in range(1, wb2_sheet.max_row+1):
    for j in range(1, wb2_sheet.max_column+1):
        wb_sheet.cell(row=i, column=j).value = wb2_sheet.cell(row=i, column=j).value
wb.save(str(today)+ ".xlsx")
os.remove(str(yesterday)+ ".xlsx")
